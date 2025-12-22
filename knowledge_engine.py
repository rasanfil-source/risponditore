import os
import logging
import openpyxl
from config import KNOWLEDGE_BASE_FILE, SHEET_LITE, SHEET_CORE, SHEET_DOCTRINE

logger = logging.getLogger(__name__)

class KnowledgeEngine:
    """Loads and provides access to the three knowledge layers.
    The content is never exposed to the user; it is used only to
    influence prompt construction.
    """
    def __init__(self, base_path: str = None):
        # If no base_path is provided, assume a 'knowledge' folder next to this file
        if base_path is None:
            base_dir = os.path.dirname(__file__)
            base_path = os.path.join(base_dir, "knowledge")
            self.excel_path = os.path.join(base_dir, KNOWLEDGE_BASE_FILE)
        else:
            self.base_path = base_path
            # Careful here: if base_path is passed, we might need to look for excel nearby or inside?
            # Assuming for now excel is in parent dir of 'knowledge' or same dir as init
            self.excel_path = os.path.join(os.path.dirname(base_path), KNOWLEDGE_BASE_FILE)

        self.base_path = base_path
        
        # Check if Excel exists
        if os.path.exists(self.excel_path):
            logger.info(f"Loading knowledge from Excel: {self.excel_path}")
            self.lite = self._load_from_excel(SHEET_LITE) or self._load("ai_core_lite.txt")
            self.core = self._load_from_excel(SHEET_CORE) or self._load("ai_core.txt")
            self.doctrine = self._load_from_excel(SHEET_DOCTRINE) or self._load("base_doctrine.txt")
        else:
            logger.warning(f"Excel KB not found at {self.excel_path}, falling back to text files")
            self.lite = self._load("ai_core_lite.txt")
            self.core = self._load("ai_core.txt")
            self.doctrine = self._load("base_doctrine.txt")

    def _load(self, filename: str) -> str:
        path = os.path.join(self.base_path, filename)
        try:
            with open(path, "r", encoding="utf-8") as f:
                return f.read()
        except FileNotFoundError:
            logger.error(f"Knowledge file missing: {filename}")
            return ""

    def _load_from_excel(self, sheet_name: str) -> str:
        try:
            wb = openpyxl.load_workbook(self.excel_path, read_only=True, data_only=True)
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in Excel KB")
                return ""
            
            ws = wb[sheet_name]
            content = []
            for row in ws.iter_rows(values_only=True):
                # Join non-None cells
                row_text = " ".join([str(cell).strip() for cell in row if cell])
                if row_text:
                    content.append(row_text)
            
            return "\n".join(content)
        except Exception as e:
            logger.error(f"Error reading sheet {sheet_name}: {e}")
            return ""

    def get_tone_guidelines(self) -> str:
        return self.lite

    def get_pastoral_guidelines(self) -> str:
        return self.core

    def get_doctrinal_content(self) -> str:
        return self.doctrine
